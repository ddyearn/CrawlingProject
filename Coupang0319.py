from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import chromedriver_autoinstaller
import subprocess
import time 

subprocess.Popen(f'google-chrome --remote-debugging-port=9222  --user-data-dir=data_dir'.split()) 
option = Options()
option.add_argument('--headless')
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
try:
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
except:
 chromedriver_autoinstaller.install(True)
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
time.sleep(5)

#login
url = 'https://login.coupang.com/login/login.pang'
driver.get(url=url)
time.sleep(5)

id_input = driver.find_element(By.XPATH, '//*[@id="login-email-input"]')
id_input.send_keys('id')

pw_input = driver.find_element(By.XPATH, '//*[@id="login-password-input"]')
pw_input.send_keys('pwd')

driver.find_element(By.XPATH, '/html/body/div[1]/div/div/form/div[5]/button').click()
driver.quit()

subprocess.Popen(f'google-chrome --remote-debugging-port=9222  --user-data-dir=data_dir'.split()) 
option = Options()
option.add_argument('--headless')
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
try:
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
except:
 chromedriver_autoinstaller.install(True)
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
time.sleep(5)


#행사
url = 'https://www.coupang.com/np/coupangbenefit'
driver.get(url=url)
time.sleep(5)

wb = Workbook()                                       
ws1  = wb.create_sheet('행사')                         
wb.remove_sheet(wb['Sheet']) 
ws1.append(['이미지', '링크'])

div = driver.find_element(By.CLASS_NAME, 'coupang-benefit-list')
list = div.find_element(By.ID, 'productList')
lis = list.find_elements(By.TAG_NAME, 'li')

for li in lis:
    img = li.find_element(By.TAG_NAME, 'img').get_attribute('src')
    url = li.find_element(By.TAG_NAME, 'a').get_attribute('href')

    print('img:' + img)
    print('URL:' + url)

    ws1.append([img, url])
driver.quit()

subprocess.Popen(f'google-chrome --remote-debugging-port=9222  --user-data-dir=data_dir'.split())
option = Options()
option.add_argument('--headless')
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
try:
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
except:
 chromedriver_autoinstaller.install(True)
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
time.sleep(5)


#광고
url = 'https://www.coupang.com/np/search?component=&q=즉석밥&channel=user'
driver.get(url=url)
time.sleep(5)

ws2 = wb.create_sheet('광고')
ws2.append(['정보'])

div = driver.find_element(By.ID, 'srpKeywordProductTopBanner')
info = div.find_elements(By.TAG_NAME, 'span')

for i in info:
    t = i.text
    print(t)
    ws2.append([t])
driver.quit()

subprocess.Popen(f'google-chrome --remote-debugging-port=9222  --user-data-dir=data_dir'.split()) 
option = Options()
option.add_argument('--headless')
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
try:
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
except:
 chromedriver_autoinstaller.install(True)
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
time.sleep(5)

#카테고리
url = 'https://www.coupang.com/np/categories/446032'
driver.get(url=url)
time.sleep(5)

ws3 = wb.create_sheet('카테고리')
ws3.append(['이름'])

list = driver.find_element(By.ID, 'searchCategoryComponent')
lis = list.find_elements(By.TAG_NAME, 'label')

for li in lis:
    name = li.text
    print(name)
    ws3.append([name])
driver.quit()

subprocess.Popen(f'google-chrome --remote-debugging-port=9222  --user-data-dir=data_dir'.split()) 
option = Options()
option.add_argument('--headless')
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
try:
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
except:
 chromedriver_autoinstaller.install(True)
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
time.sleep(5)

#일반  
url = 'https://www.coupang.com/np/search?q=top100&channel=relate'
driver.get(url=url)
time.sleep(5)

ws4 = wb.create_sheet('일반')                         
ws4.append(['이름', '가격', '배송기한', '상세페이지'])

product = driver.find_element(By.ID, 'productList')
lis = product.find_elements(By.CLASS_NAME, 'search-product')
for li in lis:
    product = li.find_element(By.CLASS_NAME, 'name').text
    price = li.find_element(By.CLASS_NAME, 'price-value').text
    delivery = li.find_element(By.CLASS_NAME, 'delivery').text
    url = li.find_element(By.CLASS_NAME, 'search-product-link').get_attribute('href')

    print('Product:' + product)
    print('Price:' + price)
    print('Delivery:' + delivery)
    print('URL:' + url)

    ws4.append([product, price, delivery, url])
driver.quit()

subprocess.Popen(f'google-chrome --remote-debugging-port=9222  --user-data-dir=data_dir'.split()) 
option = Options()
option.add_argument('--headless')
option.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]
try:
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
except:
 chromedriver_autoinstaller.install(True)
 driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver', options=option)
time.sleep(5)

#제품 상세
url = 'https://www.coupang.com/vp/products/6135394118?itemId=13195279382&vendorItemId=81917799353&q=즉석밥&itemsCount=36&searchId=c0b0d5eada62494186142bf02cd63393&rank=0&isAddedCart='
driver.get(url=url)
time.sleep(5)

ws5 = wb.create_sheet('상세')                         
ws5.append(['판매자', '상품명', '리뷰작성자', '별점', '내용', '만족도'])

driver.find_element(By.CSS_SELECTOR, '.count').click()
time.sleep(5)

div = driver.find_element(By.CLASS_NAME, 'sdp-review')
list = div.find_elements(By.TAG_NAME, 'article')
for lis in list:
    seller = lis.find_element(By.CLASS_NAME, 'sdp-review__article__list__info__product-info__seller_name').text
    user_name = lis.find_element(By.CSS_SELECTOR, '.sdp-review__article__list__info__user__name').text
    rating = lis.find_element(By.CSS_SELECTOR, '.sdp-review__article__list__info__product-info__star-orange')
    if rating == None:
        rating = 0
    else :
        rating = int(rating.get_attribute('data-rating'))
    prod_name = lis.find_element(By.CSS_SELECTOR, '.sdp-review__article__list__info__product-info__name').text
    review_content = lis.find_element(By.CSS_SELECTOR, '.sdp-review__article__list__review > div').text
    liked = lis.find_element(By.CSS_SELECTOR, '.sdp-review__article__list__survey__row__answer').text

    print(seller, prod_name, user_name, rating, review_content, liked)
    ws5.append([seller, prod_name, user_name, rating, review_content, liked])

wb.save('./coupangList.xlsx')
driver.quit()
